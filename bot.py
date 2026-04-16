import logging
import re
import os
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove, InputFile
from telegram.ext import Application, CommandHandler, CallbackQueryHandler, MessageHandler, filters, ContextTypes, ConversationHandler

TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]
CHANNEL_USERNAME = "@Alcozer_rnd"
CHANNEL_ID = "@Alcozer_rnd"
PRIVACY_POLICY_URL = "https://alcozerjewelry.ru/info/policy/"

# ⚠️ ВСТАВЬ СВОЙ TELEGRAM ID ВМЕСТО 123456789!
# Узнать ID можно у бота @userinfobot
ADMIN_ID = 1208006095  # ЗАМЕНИ НА СВОЙ ID!

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

EXCEL_FILE = os.path.join(os.path.dirname(__file__), "users_data.xlsx")

# Состояния для ConversationHandler
WAITING_FOR_FULL_NAME = 1
WAITING_FOR_PHONE = 2

def init_excel():
    if not Path(EXCEL_FILE).exists():
        wb = Workbook()
        ws = wb.active
        ws.append(["user_id", "username", "full_name", "phone", "activated_at"])
        wb.save(EXCEL_FILE)
        print("✅ Excel создан")

def user_already_activated(user_id: int) -> bool:
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] == user_id:
                return True
        return False
    except:
        return False

def save_user_data(user_id: int, username: str, full_name: str, phone: str):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([user_id, username, full_name, phone, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(EXCEL_FILE)
    print(f"✅ Сохранён пользователь {user_id}: {full_name}, {phone}")
    
    # Отправляем уведомление админу о новой активации
    try:
        # Этот код требует настройки, но пока закомментирован
        # await context.bot.send_message(chat_id=ADMIN_ID, text=f"🎉 Новая активация!\n👤 {full_name}\n📞 {phone}")
        pass
    except:
        pass

def validate_phone(phone: str) -> bool:
    """Проверяет корректность номера телефона"""
    cleaned = re.sub(r'[\s\-\(\)]', '', phone)
    
    patterns = [
        r'^\+7\d{10}$',      # +7XXXXXXXXXX
        r'^8\d{10}$',        # 8XXXXXXXXXX
        r'^7\d{10}$',        # 7XXXXXXXXXX
        r'^9\d{9}$',         # 9XXXXXXXXX
    ]
    
    for pattern in patterns:
        if re.match(pattern, cleaned):
            return True
    return False

def format_phone(phone: str) -> str:
    """Приводит номер к единому формату +7XXXXXXXXXX"""
    cleaned = re.sub(r'[\s\-\(\)]', '', phone)
    
    if cleaned.startswith('8'):
        return '+7' + cleaned[1:]
    elif cleaned.startswith('7'):
        return '+' + cleaned
    elif cleaned.startswith('9'):
        return '+7' + cleaned
    else:
        return cleaned

async def is_subscribed(user_id: int, context: ContextTypes.DEFAULT_TYPE) -> bool:
    try:
        member = await context.bot.get_chat_member(CHANNEL_ID, user_id)
        return member.status in ["member", "administrator", "creator"]
    except Exception as e:
        print(f"Ошибка проверки подписки: {e}")
        return False

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    print(f"👤 Пользователь {user_id} запустил бота")
    
    if user_already_activated(user_id):
        await update.message.reply_text("❌ Вы уже активировали сертификат.")
        return
    
    keyboard = [[InlineKeyboardButton("✅ Согласен(а) и продолжить", callback_data="agree")]]
    
    text = (
        "Добро пожаловать! Для активации подарочного сертификата NUVO необходимо принять условия.\n\n"
        f"Вы соглашаетесь с [политикой конфиденциальности]({PRIVACY_POLICY_URL}), "
        "продолжая пользоваться ботом."
    )
    
    await update.message.reply_text(
        text,
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="Markdown"
    )

async def agree(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Пользователь согласился с политикой — запрашиваем имя и фамилию"""
    query = update.callback_query
    await query.answer()
    print(f"📝 Пользователь {query.from_user.id} согласился с политикой")
    
    await query.edit_message_text(
        "📋 Пожалуйста, введите ваши ИМЯ и ФАМИЛИЮ одним сообщением:\n\n"
        "Пример: Анна Иванова\n\n"
        "Или: Анна\n\n"
        "(если напишете только имя, фамилия останется пустой)"
    )
    
    return WAITING_FOR_FULL_NAME

async def get_full_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Получаем имя и фамилию (одним сообщением)"""
    full_name_raw = update.message.text.strip()
    user_id = update.effective_user.id
    
    if len(full_name_raw) < 2:
        await update.message.reply_text(
            "❌ Слишком короткое значение. Пожалуйста, введите ваши ИМЯ и ФАМИЛИЮ:\n\n"
            "Пример: Анна Иванова"
        )
        return WAITING_FOR_FULL_NAME
    
    # Сохраняем полное имя как есть (что написал пользователь)
    context.user_data["full_name"] = full_name_raw
    print(f"📝 Пользователь {user_id} ввёл ФИО: {full_name_raw}")
    
    # Кнопки для выбора способа ввода номера
    keyboard = [
        [KeyboardButton("📞 Отправить номер автоматически", request_contact=True)],
        [KeyboardButton("✏️ Ввести номер вручную")]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, one_time_keyboard=True, resize_keyboard=True)
    
    await update.message.reply_text(
        f"✅ Спасибо, {full_name_raw}!\n\n"
        "📱 Как вы хотите поделиться номером телефона?\n\n"
        "• «Отправить номер автоматически» — номер определится сам\n"
        "• «Ввести номер вручную» — вы сами напишете номер",
        reply_markup=reply_markup
    )
    
    return WAITING_FOR_PHONE

async def handle_auto_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка автоматической отправки номера"""
    user = update.effective_user
    contact = update.message.contact
    
    if not contact or contact.user_id != user.id:
        await update.message.reply_text("❌ Ошибка, попробуйте /start")
        return ConversationHandler.END
    
    formatted_phone = format_phone(contact.phone_number)
    context.user_data["phone"] = formatted_phone
    print(f"📞 Пользователь {user.id} отправил автономер: {formatted_phone}")
    
    # Убираем клавиатуру с кнопками выбора
    await update.message.reply_text(
        f"✅ Номер принят: {formatted_phone}",
        reply_markup=ReplyKeyboardRemove()
    )
    
    # Проверяем подписку
    if await is_subscribed(user.id, context):
        await activate(update, context)
    else:
        keyboard = [[InlineKeyboardButton("✅ Проверить подписку", callback_data="check_sub")]]
        await update.message.reply_text(
            f"🔔 Вы не подписаны на канал {CHANNEL_USERNAME}\n\n"
            "Подпишитесь на канал, затем нажмите кнопку «Проверить подписку»:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    
    return ConversationHandler.END

async def handle_manual_phone_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Пользователь выбрал ручной ввод номера"""
    if update.message.text == "✏️ Ввести номер вручную":
        await update.message.reply_text(
            "📱 Введите ваш номер телефона в одном из форматов:\n\n"
            "• +7XXXXXXXXXX (например, +79123456789)\n"
            "• 8XXXXXXXXXX (например, 89123456789)\n"
            "• 7XXXXXXXXXX (например, 79123456789)\n\n"
            "Просто напишите номер цифрами:",
            reply_markup=ReplyKeyboardRemove()
        )
        return WAITING_FOR_PHONE
    
    return ConversationHandler.END

async def process_manual_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка введённого вручную номера"""
    user = update.effective_user
    phone = update.message.text.strip()
    
    # Проверяем корректность номера
    if not validate_phone(phone):
        await update.message.reply_text(
            "❌ Неверный формат номера телефона.\n\n"
            "Пожалуйста, введите номер в одном из форматов:\n"
            "• +7XXXXXXXXXX (например, +79123456789)\n"
            "• 8XXXXXXXXXX (например, 89123456789)\n"
            "• 7XXXXXXXXXX (например, 79123456789)\n\n"
            "Попробуйте ещё раз или отправьте /cancel для отмены:"
        )
        return WAITING_FOR_PHONE
    
    # Номер корректный — сохраняем
    formatted_phone = format_phone(phone)
    context.user_data["phone"] = formatted_phone
    print(f"📞 Пользователь {user.id} ввёл номер: {formatted_phone}")
    
    await update.message.reply_text(f"✅ Номер принят: {formatted_phone}")
    
    # Проверяем подписку
    if await is_subscribed(user.id, context):
        await activate(update, context)
    else:
        keyboard = [[InlineKeyboardButton("✅ Проверить подписку", callback_data="check_sub")]]
        await update.message.reply_text(
            f"🔔 Вы не подписаны на канал {CHANNEL_USERNAME}\n\n"
            "Подпишитесь на канал, затем нажмите кнопку «Проверить подписку»:",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )
    
    return ConversationHandler.END

async def cancel_input(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отмена ввода"""
    await update.message.reply_text(
        "❌ Ввод отменён. Напишите /start, чтобы начать заново.",
        reply_markup=ReplyKeyboardRemove()
    )
    return ConversationHandler.END

async def check_subscription(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Проверка подписки на канал"""
    query = update.callback_query
    await query.answer()
    user_id = query.from_user.id
    
    if await is_subscribed(user_id, context):
        await query.edit_message_text("✅ Подписка подтверждена!")
        await activate_from_callback(update, context)
    else:
        await query.edit_message_text(
            f"❌ Вы всё ещё не подписаны на канал {CHANNEL_USERNAME}\n\n"
            "👉 Подпишитесь: " + CHANNEL_USERNAME + "\n\n"
            "После подписки снова нажмите кнопку «Проверить подписку»:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("✅ Проверить подписку", callback_data="check_sub")]])
        )

async def activate_from_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Активация после проверки подписки (из callback)"""
    query = update.callback_query
    user = query.from_user
    
    if user_already_activated(user.id):
        await query.message.reply_text("❌ Сертификат уже активирован.")
        return
    
    save_user_data(
        user.id, 
        user.username or "", 
        context.user_data.get("full_name", ""),
        context.user_data.get("phone", "")
    )
    
    await query.message.reply_text(
        "✅ СЕРТИФИКАТ NUVO АКТИВИРОВАН!\n\n"
        "🎉 Покажите это сообщение консультанту в бутике NUVO.\n\n"
        "Спасибо за выбор NUVO! 💎"
    )
    print(f"🎉 Сертификат активирован для {user.id}")

async def activate(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Активация из обычного сообщения"""
    user = update.effective_user
    message = update.message
    
    if user_already_activated(user.id):
        await message.reply_text("❌ Сертификат уже активирован.")
        return
    
    save_user_data(
        user.id, 
        user.username or "", 
        context.user_data.get("full_name", ""),
        context.user_data.get("phone", "")
    )
    
    await message.reply_text(
        "✅ СЕРТИФИКАТ NUVO АКТИВИРОВАН!\n\n"
        "🎉 Покажите это сообщение консультанту в бутике NUVO.\n\n"
        "Спасибо за выбор NUVO! 💎"
    )
    print(f"🎉 Сертификат активирован для {user.id}")

async def get_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Отправляет админу Excel-файл с данными пользователей"""
    user_id = update.effective_user.id
    
    if user_id != ADMIN_ID:
        await update.message.reply_text("❌ У вас нет доступа к этой команде.")
        return
    
    try:
        if Path(EXCEL_FILE).exists():
            with open(EXCEL_FILE, 'rb') as file:
                await update.message.reply_document(
                    document=InputFile(file, filename="users_data.xlsx"),
                    caption="📊 Вот таблица с данными пользователей, активировавших сертификаты."
                )
        else:
            await update.message.reply_text("❌ Файл с данными ещё не создан.")
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка при отправке файла: {e}")

async def view_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Показывает последние активации из таблицы (без скачивания)"""
    user_id = update.effective_user.id
    
    if user_id != ADMIN_ID:
        await update.message.reply_text("❌ У вас нет доступа к этой команде.")
        return
    
    try:
        if not Path(EXCEL_FILE).exists():
            await update.message.reply_text("❌ Файл с данными ещё не создан.")
            return
        
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) <= 1:
            await update.message.reply_text("📊 Таблица пока пуста. Никто ещё не активировал сертификат.")
            return
        
        # Берём последние 5 записей (без учёта заголовка)
        last_entries = rows[-5:] if len(rows) > 5 else rows[1:]  # без заголовка
        
        message = "📊 *Последние активации:*\n\n"
        
        for row in reversed(last_entries):  # показываем от новых к старым
            if row[0] is None:  # пропускаем пустые
                continue
            name = row[2] or "—"
            phone = row[3] or "—"
            time = row[4].split()[0] if row[4] else "—"
            message += f"👤 *{name}*\n📞 {phone}\n📅 {time}\n\n"
        
        total = len(rows) - 1
        message += f"\n_📊 Всего активаций: {total}_\n"
        message += f"_💾 Чтобы скачать полную таблицу, используй /get_excel_"
        
        await update.message.reply_text(message, parse_mode="Markdown")
        
    except Exception as e:
        await update.message.reply_text(f"❌ Ошибка при чтении файла: {e}")

async def handle_any_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обработка текста вне диалога"""
    user_id = update.effective_user.id
    if user_already_activated(user_id):
        await update.message.reply_text("✅ Вы уже активировали сертификат.")
    else:
        await start(update, context)

def main():
    init_excel()
    
    app = Application.builder().token(TOKEN).build()
    
    # ConversationHandler для сбора данных
    conv_handler = ConversationHandler(
        entry_points=[CallbackQueryHandler(agree, pattern="agree")],
        states={
            WAITING_FOR_FULL_NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, get_full_name),
            ],
            WAITING_FOR_PHONE: [
                MessageHandler(filters.CONTACT, handle_auto_phone),
                MessageHandler(filters.Regex("^✏️ Ввести номер вручную$"), handle_manual_phone_start),
                MessageHandler(filters.TEXT & ~filters.COMMAND, process_manual_phone),
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel_input)],
    )
    
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("get_excel", get_excel))
    app.add_handler(CommandHandler("view_excel", view_excel))
    app.add_handler(conv_handler)
    app.add_handler(CallbackQueryHandler(check_subscription, pattern="check_sub"))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_any_text))
    
    print("🚀 Бот запущен! Напиши /start")
    app.run_polling()

if __name__ == "__main__":
    main()